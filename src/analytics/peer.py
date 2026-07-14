from __future__ import annotations

from pathlib import Path
import sqlite3

import numpy as np
import pandas as pd

from ..config import DB_PATH, REPORTS_DIR


PEER_METRICS = [
    "roe",
    "roce",
    "npm",
    "debt_equity",
    "fcf",
    "pat_cagr_5y",
    "revenue_cagr_5y",
    "eps_cagr_5y",
    "interest_coverage",
    "asset_turnover",
]


def assign_peer_group(company_row: pd.Series) -> str:
    sector = str(company_row.get("sector", "")).strip().lower()
    industry = str(company_row.get("industry", "")).strip().lower()
    if "bank" in sector or "bank" in industry:
        return "Banks"
    if "it" in sector or "software" in industry:
        return "IT Services"
    if "fmcg" in sector or "consumer" in industry:
        return "FMCG"
    if "auto" in sector or "auto" in industry:
        return "Auto"
    if "pharma" in sector or "pharma" in industry:
        return "Pharma"
    if "metal" in sector or "metal" in industry:
        return "Metals"
    if "infra" in sector or "infra" in industry:
        return "Infrastructure"
    if "energy" in sector or "oil" in industry:
        return "Energy"
    if "telecom" in sector or "telecom" in industry:
        return "Telecom"
    if "cement" in sector or "cement" in industry:
        return "Cement"
    return "Diversified"


def compute_peer_percentiles(df: pd.DataFrame, group_col: str = "peer_group", metric_cols: list[str] | None = None) -> pd.DataFrame:
    metric_cols = metric_cols or [c for c in df.columns if c not in {group_col, "ticker", "company_name"}]
    rows: list[dict[str, object]] = []
    for group, gdf in df.groupby(group_col, dropna=False):
        for metric in metric_cols:
            ordered = gdf[["ticker", metric]].dropna().copy()
            if ordered.empty:
                continue
            pct = ordered[metric].rank(pct=True, method="average")
            if metric == "debt_equity":
                pct = 1 - pct
            for ticker, value, percentile in zip(ordered["ticker"], ordered[metric], pct, strict=False):
                rows.append(
                    {
                        "ticker": ticker,
                        group_col: group,
                        "metric": metric,
                        "value": value,
                        "percentile_rank": float(percentile),
                    }
                )
    return pd.DataFrame(rows)


def build_peer_groups(companies: pd.DataFrame) -> pd.DataFrame:
    out = companies.copy()
    out["peer_group"] = out.apply(assign_peer_group, axis=1)
    return out


def load_universe_from_db(db_path: str | Path = DB_PATH) -> pd.DataFrame:
    with sqlite3.connect(db_path) as conn:
        companies = pd.read_sql_query("SELECT * FROM companies", conn)
        financials = pd.read_sql_query("SELECT * FROM financial_ratios", conn)
        base_financials = pd.read_sql_query("SELECT * FROM financials", conn)
    latest = financials.sort_values("year").groupby("ticker", as_index=False).tail(1)
    latest_fin = base_financials.sort_values("year").groupby("ticker", as_index=False).tail(1)
    merged = latest.merge(companies, on="ticker", how="left")
    merged = merged.merge(
        latest_fin[["ticker", "sales"]].rename(columns={"sales": "latest_sales"}),
        on="ticker",
        how="left",
    )
    merged["peer_group"] = merged.apply(assign_peer_group, axis=1)
    return merged


def persist_peer_percentiles(df: pd.DataFrame, db_path: str | Path = DB_PATH) -> None:
    with sqlite3.connect(db_path) as conn:
        df.to_sql("peer_percentiles", conn, if_exists="replace", index=False)


def build_peer_percentiles(db_path: str | Path = DB_PATH) -> pd.DataFrame:
    universe = load_universe_from_db(db_path)
    if "pat_cagr_5y" not in universe.columns:
        universe["pat_cagr_5y"] = np.nan
    if "revenue_cagr_5y" not in universe.columns:
        universe["revenue_cagr_5y"] = np.nan
    if "eps_cagr_5y" not in universe.columns:
        universe["eps_cagr_5y"] = np.nan
    cols = [c for c in PEER_METRICS if c in universe.columns]
    percentiles = compute_peer_percentiles(universe, metric_cols=cols)
    persist_peer_percentiles(percentiles, db_path)
    return percentiles


def peer_group_summary(df: pd.DataFrame) -> pd.DataFrame:
    numeric = [c for c in df.columns if c not in {"ticker", "company_name", "peer_group", "sector", "industry"}]
    return df.groupby("peer_group")[numeric].median(numeric_only=True).reset_index()


def export_peer_comparison(df: pd.DataFrame, percentiles: pd.DataFrame, output_path: str | Path) -> None:
    from openpyxl.styles import PatternFill

    green = PatternFill("solid", fgColor="C6E0B4")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="F4CCCC")
    gold = PatternFill("solid", fgColor="FCE4D6")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for group, group_df in df.groupby("peer_group"):
            sheet = group_df.copy()
            pct = percentiles[percentiles["peer_group"] == group].pivot(index="ticker", columns="metric", values="percentile_rank").reset_index()
            sheet = sheet.merge(pct, on="ticker", how="left")
            metric_cols = [c for c in sheet.columns if c not in {"ticker", "company_name", "peer_group"}]
            summary = sheet[metric_cols].median(numeric_only=True).to_frame().T
            summary.insert(0, "company_name", "Peer Median")
            summary.insert(0, "ticker", "")
            summary.insert(2, "peer_group", group)
            output = pd.concat([sheet, summary], ignore_index=True)
            output.to_excel(writer, sheet_name=str(group)[:31], index=False)

            ws = writer.book[str(group)[:31]]
            ws.freeze_panes = "A2"
            benchmark_ticker = sheet.iloc[0]["ticker"]
            for row_idx in range(2, ws.max_row):
                ticker = ws.cell(row=row_idx, column=1).value
                if ticker and ticker == benchmark_ticker:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = gold
                for col_idx in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col_idx).value
                    if isinstance(header, str) and header.endswith("_percentile"):
                        value = ws.cell(row=row_idx, column=col_idx).value
                        if value is None:
                            continue
                        if value >= 0.75:
                            ws.cell(row=row_idx, column=col_idx).fill = green
                        elif value <= 0.25:
                            ws.cell(row=row_idx, column=col_idx).fill = red
                        else:
                            ws.cell(row=row_idx, column=col_idx).fill = yellow


def radar_chart_path(ticker: str) -> Path:
    path = REPORTS_DIR / "radar_charts"
    path.mkdir(parents=True, exist_ok=True)
    return path / f"{ticker}_radar.png"
